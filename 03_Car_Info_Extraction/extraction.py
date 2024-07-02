'''
\mainpage The documentation for module to extract volume data and specification data.

\section Filename: extraction.py
Objective: This module is used to extract volume data and specification data from 16888 and sina website separately.
\date 9/22/2019
\author Frank Yu
\copyright
Copyright(C) 2019 Biologue Co., Ltd
8F, No 249, Dong Sec. 1, Guangming 6th Rd, Zhubei City, Hsinchu County 302 Taiwan (R.O.C.)
'''
import os
import sys
import datetime

import openpyxl
import requests
from bs4 import BeautifulSoup
from urllib.parse import urlparse
from selenium import webdriver

# To set the debug model
DEBUG = False       # False True
DEBUG_COUNTER = 10



# The url to extract volume data and specification data
url_volume = 'https://auto.16888.com/#A'
url_spec = 'http://db.auto.sina.com.cn/'

# The url to extract the mapping between country and brand
url_mapping_country = 'https://auto.16888.com/country.html'

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/45.0.2454.101 Safari/537.36', }

excel_filename = ''

# The index for volume sheet
VOLUME_INDEX = 7
VOLUME_HEADER_ROW = 2

# The constant for all months in one year
MONTH = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

# The constant for all English letters
LETTER = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

# The excel header row for volume data and specification data in excel file
excel_header_volume = ['車系', '品牌', '廠商', '型號', '級別', '價格區間(低)', '價格區間(高)']
excel_header_spec = ['車系', '品牌', '廠商', '型號', '級別', '車型', '車款', '上市時間', '廠商指導價(萬)', '電動調節-主座椅', '電動調節-副座椅', '電動調節-後排座椅', '加熱-前排座椅', '加熱-後排座椅', '通風-前排座椅', '通風-後排座椅', '按摩-前排座椅', '按摩-後排座椅', '腰部支撐-主座椅', '腰部支撐-副座椅']

# The class for car
LEVEL1_NAME = ['微型', '小型', '紧凑型', '中型', '中大型', '大型']
LEVEL2_NAME = ['两厢', '三厢', '旅行车', 'SUV', '皮卡', 'MPV', '轻客', '微卡', '微客', '跑车', '轿跑', '房车', '客车']

# The status code
STATUS_LIST = ['TIMEOUT', 'NO_DATA', 'SUCCESS']

# The counter for volume and specification
volume_counter = 0
volume_total = 0                                            # The total number of cars for volume
volume_no_data = 0                                          # The total number of cars without volume data
volume_timeout = 0                                          # The total number of cars which encounter timeout at extraction for volume
spec_counter = 0
spec_total = 0                                              # The total number of car for specification
spec_no_data = 0                                            # The total number of car without specification data
spec_timeout = 0                                            # The total number of car which encounter timeout at extraction for specification

# The variable to save the mapping between country and brand
country_mapping = {}

def show_license():
    print('**********************************************************************')
    print('*The program to extract car information of volume and specification!!*')
    print('*Author: Frank Yu                                                    *')
    print('*Company: Biologue                                                   *')
    print('*Release date: 2019/12/26                                            *')
    print('*Copyright(C) 2019 Biologue Co., Ltd                                 *')
    print('*Contact: Frank Yu                                                   *')
    print('*Email: frankyu@nitring.com                                          *')
    print('**********************************************************************')

def get_country_mapping():
    '''
    Objective: Extract the mapping between counter and brand

    @param None

    @return None
    '''
    global country_mapping
    r = requests.get(url_mapping_country)
    soup = BeautifulSoup(r.text, 'html.parser')
    countries = soup.select('.brand_title')
    for country in countries:
        country_name = country.select('.t-2')[0]['name']
        brand_tag = country.find_next_sibling()
        while brand_tag['class'][0] == 'brand_box':
            brand_name = brand_tag.select('a')[0]['name']
            country_mapping[brand_name] = country_name
            brand_tag = brand_tag.find_next_sibling()

def get_car_name_volume_url(url):
    '''
    Objective: Extract the car name and url for volume

    @param url: The url to extract the car name and url for volume

    @return All car names and urls for volume
    '''
    r = requests.get(url, headers = headers)
    soup = BeautifulSoup(r.text, 'html.parser')
    brand_boxes = soup.select('.brand_box')
    name_url = []
    for brand_box in brand_boxes:
        manufacturer = brand_box.select('.brand_list')[0].select('h1')[0].select('a')[0].text
        models = brand_box.select('.name')
        for model in models:
            # [model, manufacturer, url]
            name_url.append([model.select('a')[0].text, manufacturer, model.select('a')[0]['href']])
    return name_url

# [country, brand, manufacturer, model_name, level, low_price, high_price, volume]
def get_one_car_volume(car):
    '''
    Objective: Extract the volume for one car

    @param car: The url for this car to extract volume data

    @return The status and volume data for this car
    '''
    global volume_counter
    global volume_timeout
    global volume_no_data

    r = requests.get(car[2], headers = headers)
    volume_counter += 1
    print('volume car', volume_counter, ', progress:', '%.2f%%' % ((volume_counter / volume_total) * 100))
    if r.status_code != requests.codes.ok:
        volume_timeout += 1
        print('Error request code:', car[2])
        return STATUS_LIST[0], None
    soup = BeautifulSoup(r.text, 'html.parser')
    car_info = [car[1], car[0]]

    volume = {}
    try:
        manu = soup.select('.manufacturer')
        brand = manu[0].find_next_sibling('dd')
        brand = brand.find_all('a')
        car_info.insert(0, brand[0].text.strip())
        car_info.append(brand[1].text.strip())

        price = manu[0].select('a')[0].text
        price = price.strip()
        price = price.replace('万', '')
        price = price.split('-')
        price = [float(item) for item in price]
        if len(price) == 2:
            car_info.extend(price)
        elif len(price) == 1:
            price.append('-')
            car_info.extend(price)
    except:
        volume_no_data += 1
        return STATUS_LIST[0], None

    for key in country_mapping.keys():
        if car_info[0] in key or key in car_info[0]:
            car_info.insert(0, country_mapping[key])
            break
    else:
        car_info.insert(0, '')

    try:
        volume_links = soup.select('.menu_list')[0].select('li')
        for link in volume_links:
            link_temp = link.select('a')[0]
            if link_temp.text == '销量':
                volume_link = link_temp['href']
                break
    except:
        car_info.append(volume)
        return STATUS_LIST[1], car_info

    r = requests.get(volume_link, headers = headers)

    soup = BeautifulSoup(r.text, 'html.parser')
    volume_table = soup.select('.xl-table-def')
    volume_end = False
    monthly_volumes = volume_table[0].find_all('tr')
    for monthly_volume in monthly_volumes[1:]:
        if volume_end == True:
            break
        temp = monthly_volume.find_all('td')
        if temp[0].text == '2015-01':
            volume_end = True
        try:
            temp[1] = int(temp[1].text)
        except:
            continue
        volume[temp[0].text] = temp[1]

    while(volume_end != True):
        next_link = []
        try:
            next_link = soup.select('.xl-data-pageing')[0]
            next_link = next_link.select('a.next')
        except:
            print('volume_link:', volume_link)
        if len(next_link) == 0:
            break
        else:
            next_link = next_link[-1]
            hostname = 'https://' + urlparse(r.url).hostname
            r = requests.get(hostname + next_link['href'], headers = headers)
            soup = BeautifulSoup(r.text, 'html.parser')
            volume_table = soup.select('.xl-table-def')
            monthly_volumes = volume_table[0].find_all('tr')
            for monthly_volume in monthly_volumes[1:]:
                temp = monthly_volume.find_all('td')
                if temp[0].text == '2015-01':
                    volume_end = True
                try:
                    temp[1] = int(temp[1].text)
                except:
                    continue
                volume[temp[0].text] = temp[1]

    car_info.append(volume)
    return STATUS_LIST[2], car_info

def create_volume_xlsx_header():
    '''
    Objective: Write the header row to excel volume sheet

    @param None

    @return None
    '''
    today = datetime.date.today()

    for i in range(2015, today.year):
        for j in range(len(MONTH)):
            excel_header_volume.append(str(i) + '-' + ('%02d' % (j + 1)))

    for i in range(today.month):
        excel_header_volume.append(str(today.year) + '-' + ('%02d' % (i + 1)))

    ws1['H1'] = '銷量'
    for i in range(len(excel_header_volume[0: VOLUME_INDEX])):
        ws1.cell(row = VOLUME_HEADER_ROW, column = i + 1, value = excel_header_volume[i])
    for i in range(VOLUME_INDEX, len(excel_header_volume)):
        header_cell = MONTH[int(excel_header_volume[i][5:]) - 1] + '-' + excel_header_volume[i][2: 4]
        ws1.cell(row = VOLUME_HEADER_ROW, column = i + 1, value = header_cell)

def write_xlsx_volume(car_info):
    '''
    Objective: Write the volume data for one car to excel volume sheet

    @param car_info: The volume data for this car

    @return None
    '''
    current_row = ws1.max_row + 1
    column_index = 0
    for i in range(len(car_info[:VOLUME_INDEX])):
        ws1.cell(row = current_row, column = i + 1, value = car_info[i])
    volume = car_info[VOLUME_INDEX]
    for i in range(VOLUME_INDEX, len(excel_header_volume)):
        if len(volume) == 0:
            ws1.cell(row = current_row, column = i + 1, value = '-')
        else:
            try:
                month_volume = volume[excel_header_volume[i]]
            except KeyError:
                continue
            ws1.cell(row = current_row, column = i + 1, value = month_volume)

def get_volume_data():
    '''
    Objective: To extract all volume data from the specific websites and display the statistics

    @param None

    @return None
    '''
    create_volume_xlsx_header()
    name_url = get_car_name_volume_url(url_volume)
    global volume_total
    volume_total = len(name_url)

    for car in name_url:
        status, car_info = get_one_car_volume(car)
        if status == STATUS_LIST[0]:
            continue
        write_xlsx_volume(car_info)

        if DEBUG:
            if volume_counter > DEBUG_COUNTER:
                break

    print('Volume Total:', volume_total)
    print('Volume No data:', volume_no_data)
    print('Volume Timeout:', volume_timeout)

def create_spec_xlsx_header():
    '''
    Objective: Write the header row to excel specification sheet

    @param None

    @return None
    '''
    ws2['A1'] = '標配:1'
    ws2['B1'] = '選配:0'
    ws2['C1'] = '無:X'
    ws2.append(excel_header_spec)

def get_car_name_spec_url(url):
    '''
    Objective: Extract the car name and url for specification

    @param url: The url to extract the car name and url for specification

    @return The status and all car names and urls for volume
    '''
    name_url = []
    try:
        r = requests.get(url, headers = headers)
    except:
        print('Request error:', url)
        return False, name_url
    if not r.status_code == requests.codes.ok:
        return False, name_url

    soup = BeautifulSoup(r.text, 'html.parser')

    brand_box = soup.select('#J_scrollLeter')
    brands = brand_box[0].select('dl')
    for brand in brands:
        brand_name = brand.find_all(attrs = {'data-type': 'did'})[0].text
        brand_name = brand_name.replace('-', '·')
        manufacturers = brand.select('.maker ')
        for manufacturer in manufacturers:
            manufacturer_name = manufacturer.find(attrs = {'data-type': 'bid'}).text
            manufacturer_name = manufacturer_name.replace('-', '·')
            models = manufacturer.find_next_siblings('dd')
            for model in models:
                try:
                    model_name = model.find(attrs = {'data-type': 'subid'}).text
                    model_url = model.find(attrs = {'data-type': 'subid'})['href']
                except:
                    break
                #[brand_name, manufacturer_name, model_name, url]
                name_url.append([brand_name, manufacturer_name, model_name, model_url])

    return True, name_url

#[country, brand_name, manufacturer_name, model_name, [level1, level2, version, year, price, 電動調節, 加熱, 通風, 按摩, 腰部支撐], [level1, level2, version, year, price, 電動調節, 加熱, 通風, 按摩, 腰部支撐], ...]
def get_one_car_spec(car):
    '''
    Objective: Extract the specification for one car

    @param car: The url for this car to extract specification data

    @return The status and specification data for this car
    '''
    global spec_counter
    global spec_no_data
    global spec_timeout

    car_info = []
    r = requests.get('http:' + car[3], headers = headers)
    spec_counter += 1
    print('spec car', spec_counter, ', progress:', '%.2f%%' % ((spec_counter / spec_total) * 100))

    if r.status_code != requests.codes.ok:
        spec_timeout += 1
        print('Spec error request:', car[3])
        return STATUS_LIST[0], None
    soup = BeautifulSoup(r.text, 'html.parser')

    try:
        if soup.select('.pagelogo')[0].text.find('SINA E STATION'):
            new_power = True
        else:
            new_power = False
    except:
        new_power = False

    if new_power == True:
        try:
            price = soup.select('#main_focus')[0].select('.main_focus_r')[0].select('.strong')[0].text
        except:
            print('Get price error:', car[3])
            spec_no_data += 1
            return STATUS_LIST[0], None
    else:
        try:
            price = soup.select('.con_img')[0].select('em')[0].text
        except:
            print('Get price error:', car[3])
            spec_no_data += 1
            return STATUS_LIST[0], None

    if price == '未上市' or price == '停售':
        spec_no_data += 1
        return STATUS_LIST[0], None

    try:
        if new_power == True:
            spec_url = soup.select('#nav_estation')[0]
            spec_url = spec_url.select('a')[1]
        else:
            spec_url = soup.select('#nav')[0]
            spec_url = spec_url.select('a')[1]
    except:
        print('No link for spec:', spec_url)
        spec_no_data += 1
        return STATUS_LIST[1], None

    if spec_url.text == '配置':
        spec_url = spec_url['href']
    else:
        print('No link for spec:', spec_url)
        spec_no_data += 1
        return STATUS_LIST[0], None
    print('spec_url: ', spec_url)
    try:
        option = webdriver.ChromeOptions()
        option.add_argument('--headless')
        browser = webdriver.Chrome(options = option)
        browser.get(spec_url)
    except:
        print('Spec error request:', spec_url)
        spec_timeout += 1
        return STATUS_LIST[0], None

    soup = BeautifulSoup(browser.page_source, 'html.parser')
    browser.quit()

    try:
        prices = soup.select('#J_tableHead')[0]
        prices = prices.select('.tdsecond')
    except:
        print('No spec:', spec_url)
        spec_no_data += 1
        return STATUS_LIST[0], None

    for price_tag in prices:
        variable_version = price_tag.select('.subtit')[0].text
        year = variable_version[:4]
        try:
            year = int(year)
        except:
            year = None
        version = variable_version[5:]
        price = price_tag.select('.guide')[0].select('span')[0].text
        price = price.replace('万', '')
        try:
            price = float(price)
        except:
            print('Price conversion error:', spec_url)
            price = None
        car_info.append([version, year, price])

    try:
        all_spec = soup.select('#J_tableBody')[0]
        all_spec = all_spec.select('.con')
        all_level = all_spec[0].select('tr')[0].select('td')
    except:
        print('Get spec error:', spec_url)
        return STATUS_LIST[0], None

    levels = []
    for level_tag in all_level[1:]:
        level = level_tag.text
        level = level.split('/')
        level1 = ''
        level2 = ''
        if len(level) == 3:
            if level[2].strip() in LEVEL1_NAME:
                level1 = level[2].strip()
                level2 = ''
            elif level[2].strip() in LEVEL2_NAME:
                level2 = level[2].strip()
                level1 = ''
        elif len(level) > 3:
            if level[2].strip() in LEVEL1_NAME:
                level1 = level[2].strip()
                level2 = level[3].strip()
            if level[3].strip() in LEVEL1_NAME:
                level2 = level[2].strip()
                level1 = level[3].strip()
        levels.append([level1, level2])

    seats = []
    for i in range(len(levels)):
        car_info[i].insert(0, levels[i][1])
        car_info[i].insert(0, levels[i][0])
        seats.append([])

    seat_specs = all_spec[9].select('tr')
    # Order: 電動調節、加熱、通風、按摩、腰部支撐
    # 1: 標配， 0: 選配， X: 無
    s_temp = seat_specs[3: 7]
    s_temp.insert(0, seat_specs[1])
    for seat_spec in s_temp:
        seat_spec_tags = seat_spec.select('td')
        for i in range(1, len(seat_spec_tags)):
            seat = seat_spec_tags[i].text
            seat = seat.replace(' ', '')
            seat = seat.replace('●', '1')
            seat = seat.replace('○', '0')
            seat = seat.replace('-', 'X')
            seat = seat.split('/')
            seats[i - 1].extend(seat)

    for i in range(len(seats)):
        car_info[i].extend(seats[i])

    car_info.insert(0, car[2])
    car_info.insert(0, car[1])
    car_info.insert(0, car[0])

    for key in country_mapping.keys():
        if car[0] in key or key in car[0]:
            car_info.insert(0, country_mapping[key])
            break
    else:
        car_info.insert(0, '')

    return STATUS_LIST[2], car_info

def write_xlsx_spec(car_info):
    '''
    Objective: Write the specification data for one car to excel specification sheet

    @param car_info: The specification data for this car

    @return None
    '''
    current_row = ws2.max_row + 1
    start_column = 5
    for i in range(4, len(car_info)):
        ws2.cell(row = current_row, column = 1, value = car_info[0])
        ws2.cell(row = current_row, column = 2, value = car_info[1])
        ws2.cell(row = current_row, column = 3, value = car_info[2])
        ws2.cell(row = current_row, column = 4, value = car_info[3])
        for j in range(len(car_info[i])):
            if j >= 5:
                try:
                    car_info[i][j] = int(car_info[i][j])
                except:
                    pass
            ws2.cell(row = current_row, column = (start_column + j), value = car_info[i][j])
        current_row += 1

def get_spec_data():
    '''
    Objective: To extract all specification data from the specific websites and display the statistics

    @param None

    @return None
    '''
    global spec_total
    create_spec_xlsx_header()
    status, name_url = get_car_name_spec_url(url_spec)
    if status == False:
        return

    spec_total = len(name_url)

    for car in name_url:
        status, car_info = get_one_car_spec(car)
        if status == STATUS_LIST[0]:
            continue
        write_xlsx_spec(car_info)

        if DEBUG:
            print('spec_counter: ', spec_counter)
            if spec_counter > DEBUG_COUNTER:
                break

    print('Spec Total:', spec_total)
    print('Spec No data:', spec_no_data)
    print('Spec Timeout:', spec_timeout)

if __name__ == '__main__':
    show_license()
    # To decide if volume data and specification data are extracted
    extract_volume = input('Extract car volume: Y/N => ')
    extract_spec = input('Extract car spec: Y/N => ')
    if extract_volume == 'Y' or extract_volume == 'y' or extract_volume == 'YES':
        extract_volume = True       # False
    else:
        extract_volume = False       # True
    
    if extract_spec == 'Y' or extract_spec == 'y' or extract_spec == 'YES':
        extract_spec = True         # False
    else:
        extract_spec = False         # True

    # The excel filename to save final result
    today = datetime.datetime.today()
    excel_filename = 'Market_Research_' + ('%04d' % today.year) + ('%02d' % today.month) + ('%02d' % today.day) + ('%02d' % today.hour) + ('%02d' % today.minute) + ('%02d' % today.second) + '.xlsx'
    get_country_mapping()

    # Create excel sheet to save volume data and specification data separately.
    if os.path.exists(excel_filename):
        wb = openpyxl.load_workbook(excel_filename)
        ws1 = wb['Volume']
        ws2 = wb['Spec']
    else:
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = 'Volume'
        wb.create_sheet('Spec')
        ws2 = wb['Spec']

    # Extract volume data
    if extract_volume == True:
        get_volume_data()
        print('volume extraction done')

    # Extract specification data
    if extract_spec == True:
        get_spec_data()
        print('spec extraction done')

    # Save all data to the excel file
    wb.save(excel_filename)
    print('all extraction done')
    input('Press <enter>')